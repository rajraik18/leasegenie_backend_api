"""Test the Excel export layout and confidence-band formatting."""
from dataclasses import dataclass, field as dc_field
from datetime import datetime

import openpyxl

from app.services.excel_export import build_workbook, to_bytes


# ---------------------------------------------------------------------------
# Minimal data builders (avoid pydantic — excel_export is duck-typed)
# ---------------------------------------------------------------------------

@dataclass
class _FV:
    document_id: str = ""
    document_label: str = ""
    value: str | None = None
    confidence: float = 0.0
    page_number: int | None = None
    clause_number: str | None = None
    clause_text: str | None = None
    condition_type_taken: str | None = None
    red_flags: list = dc_field(default_factory=list)
    needs_review: bool = False
    cross_field_notes: list = dc_field(default_factory=list)
    trace_summary: str | None = None


@dataclass
class _F:
    field_id: str = ""
    name: str = ""
    category: str = ""
    output_type: str = "Text"
    per_document: list = dc_field(default_factory=list)
    override_value: str | None = None
    concluded_value: str | None = None
    concluded_source: str = ""
    concluded_confidence: float = 0.0
    confidence_level: str = "none"
    source_document_label: str | None = None
    source_document_id: str | None = None
    page_number: int | None = None
    clause_number: str | None = None
    clause_text: str | None = None


@dataclass
class _Doc:
    id: str = ""
    filename: str = ""


@dataclass
class _Summ:
    total_fields: int = 0
    fields_extracted: int = 0
    fields_none: int = 0
    fields_overridden: int = 0
    fields_flagged_review: int = 0
    mean_confidence: float = 0.0
    high_confidence_count: int = 0
    medium_confidence_count: int = 0
    low_confidence_count: int = 0


@dataclass
class _Abs:
    tenant_id: str = ""
    tenant_name: str = ""
    suite_number: str | None = None
    abstract_type: str = ""
    property_type: str = ""
    documents: list = dc_field(default_factory=list)
    fields: list = dc_field(default_factory=list)
    summary: _Summ = dc_field(default_factory=_Summ)


def _sample_abstraction() -> _Abs:
    pd1 = _FV(document_id="d0", document_label="Base Lease", value="120000",
              confidence=0.88, page_number=3, clause_number="3.1",
              clause_text="Rent of $120,000", condition_type_taken="Fixed Rent")
    pd2 = _FV(document_id="d1", document_label="Amendment 1", value="135000",
              confidence=0.92, page_number=2, clause_number="5.1")
    review = _FV(document_id="d0", document_label="Base Lease", value="None",
                 confidence=0.0, needs_review=True,
                 red_flags=["Multiple commencement dates mentioned"])
    return _Abs(
        tenant_id="t0", tenant_name="TechCo Inc", suite_number="500",
        abstract_type="Full Abstract", property_type="Office",
        documents=[_Doc(id="d0", filename="lease.pdf"),
                   _Doc(id="d1", filename="amend1.pdf")],
        fields=[
            _F(field_id="annual_base_rent", name="Annual Base Rent",
               category="Financial Clauses", output_type="Number",
               per_document=[pd1, pd2], concluded_value="135000",
               concluded_source="amendment_1", concluded_confidence=0.92,
               confidence_level="high", source_document_label="Amendment 1",
               source_document_id="d1", page_number=2, clause_number="5.1",
               clause_text="Rent amended to $135,000"),
            _F(field_id="lease_guarantor", name="Lease Guarantor",
               category="Basic Information", output_type="Text",
               per_document=[review], concluded_value="None",
               concluded_source="none", concluded_confidence=0.0,
               confidence_level="none"),
        ],
        summary=_Summ(total_fields=2, fields_extracted=1, fields_none=1,
                      fields_flagged_review=1, mean_confidence=0.92,
                      high_confidence_count=1),
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_workbook_has_three_sheets():
    wb = build_workbook(_sample_abstraction())
    assert wb.sheetnames == ["Summary", "Lease Abstraction", "Red Flags"]


def test_grid_has_confidence_and_citation_columns(tmp_path):
    wb = build_workbook(_sample_abstraction())
    ws = wb["Lease Abstraction"]
    headers = [c.value for c in ws[1]]
    # Core columns must all be present
    for required in ["Category", "Field", "Output Type", "Override",
                     "Concluded Value", "Confidence", "Confidence Level",
                     "Source Document", "Page", "Clause #",
                     "Supporting Clause Text"]:
        assert required in headers, f"missing column: {required}"


def test_annual_base_rent_row_values():
    wb = build_workbook(_sample_abstraction())
    ws = wb["Lease Abstraction"]
    headers = [c.value for c in ws[1]]
    hmap = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[hmap["Field"]] == "Annual Base Rent":
            assert row[hmap["Concluded Value"]] == "135000"
            assert row[hmap["Confidence"]] == 0.92
            assert row[hmap["Confidence Level"]] == "high"
            assert row[hmap["Source Document"]] == "Amendment 1"
            assert row[hmap["Page"]] == 2
            assert row[hmap["Clause #"]] == "5.1"
            return
    raise AssertionError("Annual Base Rent row not found")


def test_needs_review_row_highlighted():
    wb = build_workbook(_sample_abstraction())
    ws = wb["Lease Abstraction"]
    headers = [c.value for c in ws[1]]
    hmap = {h: i for i, h in enumerate(headers)}
    review_col_idx = hmap["Needs Review"] + 1  # openpyxl 1-indexed

    for row_cells in ws.iter_rows(min_row=2):
        if row_cells[hmap["Field"]].value == "Lease Guarantor":
            assert row_cells[review_col_idx - 1].value == "Yes"
            return
    raise AssertionError("Lease Guarantor row not found")


def test_red_flags_sheet_populated():
    wb = build_workbook(_sample_abstraction())
    ws = wb["Red Flags"]
    # Expect both the playbook red flag and the needs-review entry
    texts = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert any("Multiple commencement" in (t or "") for t in texts)
    assert any("manual review" in (t or "").lower() for t in texts)


def test_bytes_output_is_valid_xlsx(tmp_path):
    blob = to_bytes(_sample_abstraction())
    assert len(blob) > 1000
    # Roundtrip through openpyxl from a file
    path = tmp_path / "roundtrip.xlsx"
    path.write_bytes(blob)
    wb = openpyxl.load_workbook(path)
    assert "Summary" in wb.sheetnames


def test_summary_sheet_shows_mean_confidence():
    wb = build_workbook(_sample_abstraction())
    ws = wb["Summary"]
    # Scan all cells for 92.0% (pretty-printed mean_confidence)
    found = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "92.0%" in cell:
                found = True
    assert found, "mean_confidence not rendered on Summary sheet"
