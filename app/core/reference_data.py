"""Reference data loader.

Parses the LeaseGenie BRD spreadsheet and exposes the 72 fields, 311
questions, 987 keyword mappings, output types, and applicability matrix as
an immutable in-memory configuration. Re-uploading the BRD + restarting the
service is all it takes to update the extraction schema.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import openpyxl


ABSTRACT_TYPES = [
    "Basic Economic Abstract",
    "Financial Terms",
    "Short Form Abstract",
    "Economic",
    "Full Abstract",
]
PROPERTY_TYPES = ["Retail", "Industrial", "Office", "Mixed-Use"]


@dataclass(frozen=True)
class FieldQuestion:
    question: str
    extract: str | None
    output: str | None
    priority: int | None
    condition_type: str | None


@dataclass
class FieldConfig:
    field_id: str            # slug, e.g. "annual_base_rent"
    name: str                # "Annual Base Rent"
    category: str            # "Financial Clauses"
    output_type: str         # "Text" or "Number"
    keywords: list[str] = field(default_factory=list)
    questions: list[FieldQuestion] = field(default_factory=list)
    # applicability: abstract_type -> bool, property_type -> bool
    abstract_applicability: dict[str, bool] = field(default_factory=dict)
    property_applicability: dict[str, bool] = field(default_factory=dict)


@dataclass
class ReferenceData:
    fields: dict[str, FieldConfig]         # field_id -> FieldConfig
    fields_by_name: dict[str, str]         # lowercase name -> field_id

    def list_fields(
        self,
        abstract_type: str | None = None,
        property_type: str | None = None,
    ) -> list[FieldConfig]:
        result = []
        for f in self.fields.values():
            if abstract_type and not f.abstract_applicability.get(abstract_type, False):
                continue
            if property_type and not f.property_applicability.get(property_type, False):
                continue
            result.append(f)
        return result

    def get(self, field_id: str) -> FieldConfig | None:
        return self.fields.get(field_id)


def _slug(name: str) -> str:
    """Create a stable id from a human-readable field name."""
    out = []
    for ch in name.lower().strip():
        if ch.isalnum():
            out.append(ch)
        elif ch in (" ", "-", "_", "/"):
            out.append("_")
    s = "".join(out)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def _iter_rows(ws, min_row: int = 1) -> Iterable[tuple]:
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        yield row


def load_reference_data(brd_path: Path) -> ReferenceData:
    """Parse the BRD xlsx and build the ReferenceData object."""
    if not brd_path.exists():
        raise FileNotFoundError(f"BRD spreadsheet not found: {brd_path}")

    wb = openpyxl.load_workbook(brd_path, data_only=True)

    fields: dict[str, FieldConfig] = {}

    # --- 1. Field List: 72 fields + category + applicability ---
    ws = wb["Field List"]
    # Header row 3 has: S.No | Fields | Categorization | Basic Economic | Financial Terms |
    # Short Form | Economic | Full Abstract | Retail | Industrial | Office
    for row in _iter_rows(ws, min_row=4):
        if row is None or len(row) < 12:
            continue
        name = row[2]
        category = row[3]
        if not name or not category:
            continue
        fid = _slug(str(name))
        fc = FieldConfig(
            field_id=fid,
            name=str(name).strip(),
            category=str(category).strip(),
            output_type="Text",
        )
        fc.abstract_applicability = {
            "Basic Economic Abstract": _yesno(row[4]),
            "Financial Terms": _yesno(row[5]),
            "Short Form Abstract": _yesno(row[6]),
            "Economic": _yesno(row[7]),
            "Full Abstract": _yesno(row[8]),
        }
        fc.property_applicability = {
            "Retail": _yesno(row[9]),
            "Industrial": _yesno(row[10]),
            "Office": _yesno(row[11]),
            "Mixed-Use": True,  # Mixed-Use = union of Retail/Industrial/Office per BRD
        }
        fields[fid] = fc

    fields_by_name = {fc.name.lower(): fc.field_id for fc in fields.values()}

    # --- 2. Output_Type sheet ---
    # Some fields have multiple output_type rows in the BRD. Collect them all
    # and prefer "Number" (more specific than Text).
    output_types_seen: dict[str, set[str]] = {}
    ws = wb["Output_Type"]
    for row in _iter_rows(ws, min_row=2):
        if not row or not row[0] or not row[1]:
            continue
        fid = fields_by_name.get(str(row[0]).strip().lower())
        if fid:
            output_types_seen.setdefault(fid, set()).add(str(row[1]).strip())
    for fid, types in output_types_seen.items():
        if "Number" in types:
            fields[fid].output_type = "Number"
        elif types:
            fields[fid].output_type = next(iter(types))

    # --- 3. Keywords_Mapping sheet (multiple rows per field allowed) ---
    ws = wb["Keywords_Mapping"]
    for row in _iter_rows(ws, min_row=2):
        if not row or not row[0] or not row[1]:
            continue
        fid = fields_by_name.get(str(row[0]).strip().lower())
        if fid:
            kw = str(row[1]).strip()
            if kw and kw not in fields[fid].keywords:
                fields[fid].keywords.append(kw)

    # --- 4. Questions sheet ---
    ws = wb["Questions"]
    for row in _iter_rows(ws, min_row=4):
        if not row or len(row) < 7:
            continue
        field_name = row[1]
        question = row[4]
        if not field_name or not question:
            continue
        fid = fields_by_name.get(str(field_name).strip().lower())
        if not fid:
            continue
        try:
            priority = int(row[3]) if row[3] not in (None, "") else None
        except (TypeError, ValueError):
            priority = None
        q = FieldQuestion(
            question=str(question).strip(),
            extract=str(row[5]).strip() if row[5] else None,
            output=str(row[6]).strip() if row[6] else None,
            priority=priority,
            condition_type=str(row[2]).strip() if row[2] else None,
        )
        fields[fid].questions.append(q)

    # Sort questions per field by priority (None last)
    for fc in fields.values():
        fc.questions.sort(key=lambda q: (q.priority is None, q.priority or 0))

    return ReferenceData(fields=fields, fields_by_name=fields_by_name)


def _yesno(v) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("yes", "y", "true", "1")


@lru_cache(maxsize=1)
def get_reference_data() -> ReferenceData:
    """Cached singleton accessor."""
    from app.config import settings
    return load_reference_data(settings.brd_path)
