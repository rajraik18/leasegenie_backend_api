"""Excel export of a tenant abstraction.

Produces an .xlsx file matching the BRD's 'Lease Abstraction' sheet layout:

    Tenant header block (tenant name, suite, property, abstract type)
    Summary block (field counts, confidence rollups)

    Grid: rows = fields (grouped by category), columns:
        Category | Field | Base Lease | Amendment 1..7 | Override | Concluded Value
                                                                    | Confidence | Source | Page | Clause | Clause Text

    Red flags sheet
    Audit log sheet

Follows the xlsx skill guidance:
  * Arial 10 for body, 11 bold for headers, 14 bold for titles
  * Frozen header row + frozen first two columns
  * Confidence formatted as 0.0% (and coloured: high=green, medium=amber, low=red)
  * Conditional formatting: "None" values dimmed, needs_review rows highlighted
"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.schemas.models import TenantAbstractionOut

logger = logging.getLogger(__name__)


# Style constants
FONT_BODY = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="1F2937")

FILL_HEADER = PatternFill("solid", start_color="1F2937")       # slate
FILL_SECTION = PatternFill("solid", start_color="E5E7EB")      # light gray
FILL_HIGH = PatternFill("solid", start_color="DCFCE7")         # green tint
FILL_MEDIUM = PatternFill("solid", start_color="FEF3C7")       # amber tint
FILL_LOW = PatternFill("solid", start_color="FEE2E2")          # red tint
FILL_REVIEW = PatternFill("solid", start_color="FED7AA")       # orange tint

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


CATEGORY_ORDER = [
    "Basic Information",
    "Financial Clauses",
    "Reimbursements",
    "Critical Clauses",
    "Other Lease Clauses",
]


def build_workbook(abstraction: TenantAbstractionOut) -> Workbook:
    """Build the abstraction workbook in memory."""
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_header(wb, abstraction)
    _sheet_abstraction(wb, abstraction)
    _sheet_red_flags(wb, abstraction)

    return wb


def write_to_path(abstraction: TenantAbstractionOut, out_path: Path) -> Path:
    wb = build_workbook(abstraction)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    logger.info("Exported abstraction to %s", out_path)
    return out_path


def to_bytes(abstraction: TenantAbstractionOut) -> bytes:
    """Return the workbook as an in-memory bytes blob."""
    wb = build_workbook(abstraction)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Sheet 1 — Header + Summary
# ---------------------------------------------------------------------------

def _sheet_header(wb: Workbook, ab: TenantAbstractionOut) -> None:
    ws = wb.create_sheet("Summary")

    ws["A1"] = "Lease Abstraction Report"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="6B7280")
    ws.merge_cells("A2:F2")

    # Tenant block
    ws["A4"] = "Tenant"
    ws["A4"].font = FONT_SECTION
    ws["A4"].fill = FILL_SECTION
    rows = [
        ("Tenant Name", ab.tenant_name),
        ("Suite", ab.suite_number or ""),
        ("Abstract Type", ab.abstract_type),
        ("Property Type", ab.property_type),
        ("Documents", ", ".join(d.filename for d in ab.documents)),
    ]
    for i, (label, value) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=i, column=2, value=str(value))
        c.font = FONT_BODY
        c.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    # Summary block
    row = 5 + len(rows) + 1
    ws.cell(row=row, column=1, value="Extraction Summary").font = FONT_SECTION
    ws.cell(row=row, column=1).fill = FILL_SECTION
    row += 1
    s = ab.summary
    summary_rows = [
        ("Total Fields", s.total_fields, None),
        ("Fields Extracted", s.fields_extracted, None),
        ("Fields = 'None'", s.fields_none, None),
        ("Fields Overridden", s.fields_overridden, None),
        ("Flagged for Review", s.fields_flagged_review, FILL_REVIEW if s.fields_flagged_review > 0 else None),
        ("Mean Confidence", f"{s.mean_confidence:.1%}", None),
        ("High Confidence (≥80%)", s.high_confidence_count, FILL_HIGH),
        ("Medium Confidence (50–80%)", s.medium_confidence_count, FILL_MEDIUM),
        ("Low Confidence (<50%)", s.low_confidence_count, FILL_LOW if s.low_confidence_count > 0 else None),
    ]
    for label, value, fill in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.font = FONT_BODY
        if fill is not None:
            c.fill = fill
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    for letter in ("C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 20


# ---------------------------------------------------------------------------
# Sheet 2 — Main abstraction grid
# ---------------------------------------------------------------------------

def _sheet_abstraction(wb: Workbook, ab: TenantAbstractionOut) -> None:
    ws = wb.create_sheet("Lease Abstraction")

    # Determine max amendment columns from documents
    doc_labels = [d.filename for d in ab.documents]  # preserved for display
    # Per-doc columns use document_label from per_document values — more predictable
    per_doc_labels: list[str] = []
    seen = set()
    for f in ab.fields:
        for pd in f.per_document:
            if pd.document_label not in seen:
                seen.add(pd.document_label)
                per_doc_labels.append(pd.document_label)
    # Order: Base Lease first, then Amendment 1..7
    per_doc_labels.sort(key=lambda s: (0 if s == "Base Lease" else int(s.replace("Amendment ", "")) if s.startswith("Amendment ") else 99))

    # Build header row
    headers = ["Category", "Field", "Output Type"]
    headers.extend(per_doc_labels)
    headers.extend([
        "Override", "Concluded Value", "Confidence", "Confidence Level",
        "Source Document", "Page", "Clause #", "Supporting Clause Text",
        "Condition Type", "Red Flags", "Needs Review", "Cross-field Notes",
    ])

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    ws.freeze_panes = "C2"   # freeze header + first 2 cols

    # Order fields by BRD category order, then by name
    ordered = sorted(
        ab.fields,
        key=lambda f: (CATEGORY_ORDER.index(f.category) if f.category in CATEGORY_ORDER else 99, f.name),
    )

    row = 2
    for f in ordered:
        ws.cell(row=row, column=1, value=f.category).font = FONT_BODY
        ws.cell(row=row, column=2, value=f.name).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=3, value=f.output_type).font = FONT_BODY

        # per-doc values
        col_map = {lbl: i for i, lbl in enumerate(per_doc_labels, start=4)}
        for pd in f.per_document:
            col = col_map.get(pd.document_label)
            if col:
                cell = ws.cell(row=row, column=col, value=pd.value or "None")
                cell.font = FONT_BODY
                cell.alignment = ALIGN_WRAP
                # Dim None values
                if (pd.value or "None").lower() == "none":
                    cell.font = Font(name="Arial", size=10, italic=True, color="9CA3AF")

        base_col = 4 + len(per_doc_labels)
        ws.cell(row=row, column=base_col + 0, value=f.override_value or "").font = FONT_BODY
        ws.cell(row=row, column=base_col + 1, value=f.concluded_value or "None").font = Font(name="Arial", size=10, bold=True)

        # Confidence + level with colour band
        conf_cell = ws.cell(row=row, column=base_col + 2, value=f.concluded_confidence)
        conf_cell.number_format = "0.0%"
        conf_cell.font = FONT_BODY

        level_cell = ws.cell(row=row, column=base_col + 3, value=f.confidence_level)
        level_cell.font = FONT_BODY
        level_cell.alignment = ALIGN_CENTER
        if f.confidence_level == "high":
            level_cell.fill = FILL_HIGH
            conf_cell.fill = FILL_HIGH
        elif f.confidence_level == "medium":
            level_cell.fill = FILL_MEDIUM
            conf_cell.fill = FILL_MEDIUM
        elif f.confidence_level == "low":
            level_cell.fill = FILL_LOW
            conf_cell.fill = FILL_LOW

        ws.cell(row=row, column=base_col + 4, value=f.source_document_label or "").font = FONT_BODY
        ws.cell(row=row, column=base_col + 5, value=f.page_number if f.page_number is not None else "").font = FONT_BODY
        ws.cell(row=row, column=base_col + 6, value=f.clause_number or "").font = FONT_BODY

        clause_cell = ws.cell(row=row, column=base_col + 7,
                               value=(f.clause_text or "")[:1500])
        clause_cell.font = Font(name="Arial", size=9, color="374151")
        clause_cell.alignment = ALIGN_WRAP

        # Aggregate per-doc metadata for the row
        conds = {pd.condition_type_taken for pd in f.per_document if pd.condition_type_taken}
        rflags = {rf for pd in f.per_document for rf in pd.red_flags}
        needs_review = any(pd.needs_review for pd in f.per_document)
        cross_notes = {n for pd in f.per_document for n in pd.cross_field_notes}

        ws.cell(row=row, column=base_col + 8, value=", ".join(sorted(conds))).font = FONT_BODY
        ws.cell(row=row, column=base_col + 9, value="; ".join(sorted(rflags))).font = FONT_BODY

        nr = ws.cell(row=row, column=base_col + 10, value="Yes" if needs_review else "")
        nr.font = FONT_BODY
        if needs_review:
            nr.fill = FILL_REVIEW
            for col_idx in range(1, base_col + 12):
                ws.cell(row=row, column=col_idx).fill = FILL_REVIEW

        ws.cell(row=row, column=base_col + 11, value="; ".join(sorted(cross_notes))).font = FONT_BODY

        # Row borders
        for col_idx in range(1, base_col + 12):
            if not ws.cell(row=row, column=col_idx).border.left.style:
                ws.cell(row=row, column=col_idx).border = THIN_BORDER

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22    # Category
    ws.column_dimensions["B"].width = 30    # Field
    ws.column_dimensions["C"].width = 12    # Output Type
    for i, _ in enumerate(per_doc_labels, start=4):
        ws.column_dimensions[get_column_letter(i)].width = 24
    # trailing columns
    tail_widths = [20, 26, 12, 14, 20, 8, 12, 60, 22, 30, 14, 32]
    for offset, width in enumerate(tail_widths):
        col = 4 + len(per_doc_labels) + offset
        ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Sheet 3 — Red flags
# ---------------------------------------------------------------------------

def _sheet_red_flags(wb: Workbook, ab: TenantAbstractionOut) -> None:
    ws = wb.create_sheet("Red Flags")
    headers = ["Severity", "Field", "Source", "Red Flag"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    row = 2
    for f in ab.fields:
        for pd in f.per_document:
            for rf in pd.red_flags:
                ws.cell(row=row, column=1, value="info").font = FONT_BODY
                ws.cell(row=row, column=2, value=f.name).font = FONT_BODY
                ws.cell(row=row, column=3, value=pd.document_label).font = FONT_BODY
                ws.cell(row=row, column=4, value=rf).font = FONT_BODY
                ws.cell(row=row, column=1).fill = FILL_LOW
                row += 1
            if pd.needs_review:
                ws.cell(row=row, column=1, value="review").font = FONT_BODY
                ws.cell(row=row, column=2, value=f.name).font = FONT_BODY
                ws.cell(row=row, column=3, value=pd.document_label).font = FONT_BODY
                ws.cell(row=row, column=4, value="Playbook flagged for manual review").font = FONT_BODY
                ws.cell(row=row, column=1).fill = FILL_REVIEW
                row += 1

    if row == 2:
        ws.cell(row=row, column=1, value="(no red flags)").font = Font(name="Arial", italic=True, color="9CA3AF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 80
    ws.freeze_panes = "A2"
