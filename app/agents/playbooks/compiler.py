"""Playbook compiler.

Parses the 5 BRD .docx abstraction guides and Questions.xlsx, then emits one
JSON playbook per lease field into `data/playbooks_compiled/`.

Run from the project root:
    python -m app.agents.playbooks.compiler

This reads:
    data/playbooks_source/BASIC_INFORMATION.docx
    data/playbooks_source/FINANCIAL_TERMS.docx
    data/playbooks_source/REIMBURSEMENTS.docx
    data/playbooks_source/CRITICAL_LEASE_CLAUSES.docx
    data/playbooks_source/OTHER_LEASE_CLAUSES.docx
    data/playbooks_source/Questions.xlsx

And writes:
    data/playbooks_compiled/<field_id>.json   — one file per field
    data/playbooks_compiled/_index.json       — field_id → metadata

Parsing strategy:
    1. For each .docx, walk the paragraph stream. Use Heading3 = "SECTION 1/2/3"
       for doc-level structure. Use Heading4 paragraphs to detect field
       boundaries within SECTION 2 — QUESTIONS.
    2. Inside a field block, collect:
         - General info / overview text
         - Every QUESTION N: ... line (stop at the next Heading4)
         - Every "IF YES → ..." and "IF NO → ..." branch line that follows
    3. Merge with the structured rows from Questions.xlsx (which supply
       Condition Type, Priority, Output, Red Flag) by matching field name.
    4. Merge keywords per field from the Keywords_Mapping sheet.
    5. Detect property-type restrictions ("in retail leases only", etc.).
    6. Normalize field names to BRD slugs (matches app.core.reference_data).
    7. Emit one Playbook JSON per field.
"""
from __future__ import annotations

import json
import re
import sys
import zipfile
from dataclasses import asdict
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

from app.agents.playbooks.schema import (
    ActionType, Playbook, PlaybookAction, PlaybookQuestion, SearchScope,
)


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W_NS}


# ---------------------------------------------------------------------------
# Field-name → canonical BRD slug mapping (matches app.core.reference_data._slug)
# ---------------------------------------------------------------------------

def _slug(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", name.lower().strip())
    return s.strip("_")


# Docx heading text → canonical BRD field name (from Field List tab of BRD xlsx)
DOCX_FIELD_ALIASES: dict[str, str] = {
    # BASIC_INFORMATION.docx
    "Landlord (LL)": "Landlord Name",
    "Tenant (T)": "Tenant Name",
    "Lease Date": "Lease Date",
    "Building": "Building",
    "Suite #": "Suite",
    "Lease Guarantor": "Lease Guarantor",
    "Property Name": "Property name",
    "Street Address": "Street Address",
    "City": "City",
    "State": "State",
    "Lease Commencement Date": "Original Lease Commencement Date",
    "Lease Expiration Date": "Lease Expiration Date",
    "Term Commencement Date": "Term Commencement date",
    "Rent Commencement Date": "Rent Commencement Date",
    "Lease Term": "Lease Term (yrs.)",
    "Square Footage": "Leased RSF",
    "Most Recent Lease Start Date": "Most Recent Lease Start",

    # FINANCIAL_TERMS.docx
    "Annual Rent": "Annual Base Rent",
    "Future Rent Step": "Future Rent Steps",
    "Percentage Rent and Breakpoint": "Percentage Rent",
    "Security Deposit": "Security Deposit",

    # REIMBURSEMENTS.docx
    "Common Area Maintenance (CAM)": "CAM",
    "Real Estate Taxes (RE Taxes)": "RE Taxes",
    "Landlord Insurance": "Landlord Insurance",
    "Other Income, Advertising and Marketing": "Other Income (Exterior Signage/ Storage)",
    "Tenant Insurance Requirements": "Tenant Insurance Requirements",
    "Utilities": "Utilities",

    # CRITICAL_LEASE_CLAUSES.docx
    "Continuous Operations": "Continuous Operation",
    "Contraction Option": "Contraction Option",
    "Co-Tenancy (in retail leases only)": "Co-Tenancy",
    "Exclusive Use (in retail lease only)": "Exclusive Use",
    "Go-Dark": "Go-Dark",
    "Landlord Termination": "Landlord Termination",
    "Landlord Recapture Right": "Landlord's Recapture Rights",
    "Permitted Use": "Permitted Use",
    "Purchase Option": "Purchase Option",
    "Relocation": "Relocation",
    "Renewal Options": "Renewal Options",
    "Right of Expansion": "Right of Expansion",
    "ROFO": "ROFO",
    "ROFR (in office leases only)": "ROFR",
    "Sales Kick-Out (in retail leases only)": "Sales Kick-Out",
    "Tenant Termination": "Tenant Termination",

    # OTHER_LEASE_CLAUSES.docx
    "Allowance": "Allowance",
    "Alteration": "Alteration",
    "Assignment and Subletting": "Assignment and Subletting",
    "Casualty": "Casualty",
    "Condemnation": "Condemnation",
    "Hazardous Materials": "Hazardous Materials",
    "Holdover": "Holdover",
    "Landlord Restriction": "Landlord Restriction",
    "Monetary Default": "Monetary Default",
    "Non – Monetary Default": "Non-Monetary Default",
    "Parking": "Parking",
    "Repair and Maintenance": "Repair and Maintenance",
    "Reporting of Financial Information": "Reporting of Financial Information",
    "Reporting of Gross Sales": "Reporting of Gross Sales",
    "Sublease Provision": "Sublease Provision",
    "Subordination": "Subordination",
}


# Docx file → default category name for fields in that file
DOCX_CATEGORY: dict[str, str] = {
    "BASIC_INFORMATION.docx": "Basic Information",
    "FINANCIAL_TERMS.docx": "Financial Clauses",
    "REIMBURSEMENTS.docx": "Reimbursements",
    "CRITICAL_LEASE_CLAUSES.docx": "Critical Clauses",
    "OTHER_LEASE_CLAUSES.docx": "Other Lease Clauses",
}


# Questions.xlsx "Main Fields" → canonical BRD category
QXLSX_CATEGORY_MAP: dict[str, str] = {
    "Basic Information": "Basic Information",
    "Financial Terms": "Financial Clauses",
    "Financial Clauses": "Financial Clauses",
    "Reimbursement": "Reimbursements",
    "Reimbursements": "Reimbursements",
    "Critical Clauses": "Critical Clauses",
    "Other Lease Clauses": "Other Lease Clauses",
}


# Field-name variants used by Questions.xlsx that don't exactly match BRD Field List
QXLSX_FIELD_ALIASES: dict[str, str] = {
    "Suite": "Suite",
    "Original Lease Commencement Date (OLCD)": "Original Lease Commencement Date",
    "Lease Expiration Date (LED)": "Lease Expiration Date",
    "Rent Commencement Date (RCD)": "Rent Commencement Date",
    "Term Commencement Date (TCD)": "Term Commencement date",
    "Lease Term (yrs.)": "Lease Term (yrs.)",
    "Leased RSF": "Leased RSF",
    "Most Recent Lease Start": "Most Recent Lease Start",
    "Percentage Rent": "Percentage Rent",
    "Breakpoint": "Breakpoint",
    "Property Type": "Property name",
}


# ---------------------------------------------------------------------------
# Docx paragraph iterator
# ---------------------------------------------------------------------------

def _iter_docx(path: Path):
    """Yield (style, content). content is str for paragraphs, list[list[str]] for tables."""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8")
    root = ET.fromstring(xml)
    body = root.find("w:body", NS)
    for child in body:
        tag = child.tag.split("}")[-1]
        if tag == "p":
            text = "".join(t.text or "" for t in child.iter(f"{{{W_NS}}}t"))
            pPr = child.find("w:pPr", NS)
            style = ""
            if pPr is not None:
                pStyle = pPr.find("w:pStyle", NS)
                if pStyle is not None:
                    style = pStyle.get(f"{{{W_NS}}}val", "")
            yield style, text.strip()
        elif tag == "tbl":
            rows = []
            for row in child.findall("w:tr", NS):
                cells = ["".join(tx.text or "" for tx in cell.iter(f"{{{W_NS}}}t")).strip()
                        for cell in row.findall("w:tc", NS)]
                rows.append(cells)
            yield "TABLE", rows


# ---------------------------------------------------------------------------
# Section 2 extractor — pulls (field_heading → list of raw lines) from a docx
# ---------------------------------------------------------------------------

def _extract_field_blocks(docx_path: Path) -> dict[str, dict]:
    """Return { docx_field_name: {overview, question_lines, tables, preliminary} }."""
    blocks: dict[str, dict] = {}
    current_field: str | None = None
    current_block: dict | None = None
    in_section2 = False
    preliminary_lines: list[str] = []
    preamble_mode = True

    for style, content in _iter_docx(docx_path):
        # Preamble (before SECTION 1) — "Preliminary Instruction:" lines
        if preamble_mode and isinstance(content, str):
            if "SECTION 1" in content or style == "Heading3":
                preamble_mode = False
            else:
                if content and content not in ("", "Preliminary Instruction: Search for the entire lease including:"):
                    preliminary_lines.append(content)

        # Section 2 marker
        if isinstance(content, str) and "SECTION 2" in content and "QUESTIONS" in content.upper():
            in_section2 = True
            continue
        if isinstance(content, str) and "SECTION 3" in content:
            in_section2 = False
            # stash last block
            if current_field and current_block is not None:
                blocks[current_field] = current_block
            current_field = None
            current_block = None
            continue

        if not in_section2:
            continue

        # Field boundary = Heading4
        if style == "Heading4" and isinstance(content, str) and content:
            if current_field and current_block is not None:
                blocks[current_field] = current_block
            current_field = content
            current_block = {
                "overview": [],
                "question_lines": [],
                "tables": [],
            }
            continue

        if current_block is None:
            continue

        if style == "TABLE":
            current_block["tables"].append(content)
        elif isinstance(content, str) and content:
            # Classify line
            if re.match(r"^(QUESTION|Question)\s*\d+\s*[:\.]", content, re.IGNORECASE):
                current_block["question_lines"].append(("QUESTION", content))
            elif content.upper().startswith(("IF YES", "IF NO", "IF YES ", "IF NO ",
                                             "IF YES→", "IF NO→", "IF YES ->", "IF NO ->",
                                             "IFYES", "IFNO")):
                current_block["question_lines"].append(("BRANCH", content))
            elif re.match(r"^If\s+(Yes|No)", content, re.IGNORECASE):
                current_block["question_lines"].append(("BRANCH", content))
            elif content.startswith("Note:") or content.startswith("Please note"):
                current_block["question_lines"].append(("NOTE", content))
            elif content.lower().startswith("key questions"):
                current_block["question_lines"].append(("HEADING", content))
            else:
                current_block["overview"].append(content)

    if current_field and current_block is not None:
        blocks[current_field] = current_block

    # stash preliminary on a synthetic key
    blocks["_preliminary"] = {"lines": preliminary_lines}
    return blocks


# ---------------------------------------------------------------------------
# Question-line parser — turn "QUESTION 1: ...\nIF YES → X\nIF NO → Y" into
# a structured list of (qid, text, yes_branch_text, no_branch_text, notes)
# ---------------------------------------------------------------------------

_QUESTION_RE = re.compile(r"^(?:QUESTION|Question)\s*(\d+)\s*[:\.]\s*(.+)", re.IGNORECASE)
_BRANCH_RE = re.compile(r"^(If\s+YES|If\s+NO|IF\s+YES|IF\s+NO)\s*[→\->:]*\s*(.+)", re.IGNORECASE)
_GOTO_RE = re.compile(r"(?:Go to|go to|goto)\s*(?:QUESTION|Question|Q)\s*(\d+)", re.IGNORECASE)


def _parse_question_stream(lines: list[tuple[str, str]]) -> list[dict]:
    """Group lines into one record per QUESTION."""
    records: list[dict] = []
    current: dict | None = None
    for kind, text in lines:
        if kind == "QUESTION":
            if current:
                records.append(current)
            m = _QUESTION_RE.match(text)
            if not m:
                continue
            current = {
                "qid": f"Q{m.group(1)}",
                "text": m.group(2).strip(),
                "yes": None,
                "no": None,
                "notes": [],
            }
        elif kind == "BRANCH" and current is not None:
            m = _BRANCH_RE.match(text)
            if not m:
                continue
            label = m.group(1).upper().replace(" ", "")
            body = m.group(2).strip()
            if "YES" in label:
                current["yes"] = body
            elif "NO" in label:
                current["no"] = body
        elif kind == "NOTE" and current is not None:
            current["notes"].append(text)
    if current:
        records.append(current)
    return records


def _action_from_branch_text(text: str | None) -> PlaybookAction | None:
    """Convert human 'Go to QUESTION 3' / 'Extract amount' / 'Mark for review' / 'None'
    into a PlaybookAction."""
    if not text:
        return None
    low = text.lower().strip()

    goto_match = _GOTO_RE.search(text)

    # "Extract ... go to Q3"  /  "Extract the details and go to QUESTION 3."
    extract_keywords = ("extract", "abstract", "update in", "capture", "record")
    has_extract = any(kw in low for kw in extract_keywords) and not low.startswith(("if no", "if yes"))

    # Manual review signals
    if "manual review" in low or "mark" in low and "review" in low or "highlight" in low and "review" in low:
        return PlaybookAction.flag()

    # Literal fallback
    if '"undated"' in low or "record 'undated'" in low or "undated" in low and "record" in low:
        return PlaybookAction.record("undated")
    if '"no recovery"' in low or "no recovery" in low and "extract" in low:
        return PlaybookAction.record("No Recovery")
    if low in ("none", 'extract "none".', 'extract "none"', "record none", "update none", 'update "none"'):
        return PlaybookAction.record_none()
    if low.startswith("extract \"none\"") or low.startswith("extract 'none'"):
        return PlaybookAction.record_none()
    if "extract" in low and ("none" in low or "\"none\"" in low or "'none'" in low):
        return PlaybookAction.record_none()

    # Goto + extract
    if goto_match and has_extract:
        return PlaybookAction.extract_then(f"Q{goto_match.group(1)}")

    # Pure goto
    if goto_match:
        return PlaybookAction.goto_q(f"Q{goto_match.group(1)}")

    # Extract and finalize (no goto)
    if has_extract:
        return PlaybookAction.extract_then(None)  # extract + terminate

    # Fallback — treat as finalize if unclassifiable
    return PlaybookAction.finalize()


# ---------------------------------------------------------------------------
# Questions.xlsx loader — the structured metadata
# ---------------------------------------------------------------------------

def _load_questions_xlsx(path: Path) -> dict:
    """Returns {
        'questions_by_field': { canonical_field_name: [ {condition_type, priority, question, extract, output, red_flag}, ... ] },
        'keywords_by_field': { canonical_field_name: [str, ...] },
        'output_type_by_field': { canonical_field_name: str },
        'property_applicability': { canonical_field_name: {retail: bool, ...} },
    }"""
    wb = openpyxl.load_workbook(path, data_only=True)

    questions_by_field: dict[str, list[dict]] = {}
    ws = wb["Questions"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 11 or not row[1]:
            continue
        field_raw = str(row[1]).strip()
        field_name = QXLSX_FIELD_ALIASES.get(field_raw, field_raw)
        try:
            prio = int(row[3]) if row[3] not in (None, "") else None
        except (TypeError, ValueError):
            prio = None
        rec = {
            "main_field": str(row[0]).strip() if row[0] else None,
            "condition_type": str(row[2]).strip() if row[2] else "",
            "priority": prio,
            "question": str(row[4]).strip() if row[4] else "",
            "extract": str(row[5]).strip() if row[5] else "",
            "output": str(row[6]).strip() if row[6] else "",
            "red_flag": str(row[10]).strip() if row[10] else None,
        }
        questions_by_field.setdefault(field_name, []).append(rec)

    # sort by priority
    for qs in questions_by_field.values():
        qs.sort(key=lambda q: (q["priority"] is None, q["priority"] or 0))

    # Keywords
    keywords_by_field: dict[str, list[str]] = {}
    ws = wb["Keywords_Mapping"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        field_name = str(row[0]).strip()
        kw = str(row[1]).strip()
        if kw:
            keywords_by_field.setdefault(field_name, []).append(kw)

    # Output_Type
    output_by_field: dict[str, set[str]] = {}
    ws = wb["Output_Type"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        output_by_field.setdefault(str(row[0]).strip(), set()).add(str(row[1]).strip())
    output_type_by_field = {
        k: ("Number" if "Number" in v else next(iter(v)))
        for k, v in output_by_field.items()
    }

    # Property_Type
    prop_by_field: dict[str, dict[str, bool]] = {}
    ws = wb["Property_Type"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        fn = str(row[0]).strip()
        pt = str(row[1]).strip()
        prop_by_field.setdefault(fn, {})[pt] = True

    return {
        "questions_by_field": questions_by_field,
        "keywords_by_field": keywords_by_field,
        "output_type_by_field": output_type_by_field,
        "property_applicability": prop_by_field,
    }


# ---------------------------------------------------------------------------
# Merge docx tree + Questions.xlsx → Playbook
# ---------------------------------------------------------------------------

def _detect_property_restriction(docx_field_name: str) -> dict[str, bool] | None:
    """Extract explicit property restriction from the H4 heading, e.g. 'Co-Tenancy (in retail leases only)'."""
    low = docx_field_name.lower()
    if "retail lease" in low or "retail only" in low:
        return {"Retail": True, "Industrial": False, "Office": False, "Mixed-Use": True}
    if "office lease" in low or "office only" in low:
        return {"Retail": False, "Industrial": False, "Office": True, "Mixed-Use": True}
    if "industrial only" in low:
        return {"Retail": False, "Industrial": True, "Office": False, "Mixed-Use": True}
    return None


def _detect_search_scope(question_text: str) -> SearchScope:
    q = question_text.lower()
    if "summary" in q and ("page" in q or "section" in q):
        return SearchScope.SUMMARY
    if "definition" in q:
        return SearchScope.DEFINITIONS
    if "body of the lease" in q or "main lease" in q or "main section" in q or "main clause" in q:
        return SearchScope.BODY
    if "amendment" in q:
        return SearchScope.AMENDMENTS
    return SearchScope.ALL


def _normalize_output(output_str: str | None, fallback: str) -> str:
    if not output_str:
        return fallback
    s = output_str.strip()
    # Keep the richer-than-simple when present (e.g. "Currency and Text")
    return s


def _compile_one_playbook(
    docx_field_name: str,
    docx_block: dict,
    category: str,
    source_docx: str,
    qx: dict,
    preliminary: list[str],
) -> Playbook | None:
    """Build one Playbook by merging docx flow + Questions.xlsx metadata."""

    canonical_name = DOCX_FIELD_ALIASES.get(docx_field_name)
    if canonical_name is None:
        # skip unknown fields — they're not part of the 72-field BRD list
        return None

    field_id = _slug(canonical_name)

    # Merge question metadata from xlsx
    xlsx_qs = qx["questions_by_field"].get(canonical_name, [])

    # Parse the docx flow to get YES/NO branches
    docx_records = _parse_question_stream(docx_block.get("question_lines", []))

    # Index xlsx by priority — each docx QUESTION corresponds roughly to the xlsx
    # row of the same priority. This is heuristic but the BRD numbers match in
    # most cases; when they don't, we still keep the docx flow as the source of
    # truth for branching, and overlay xlsx metadata where priorities line up.
    xlsx_by_prio = {q["priority"]: q for q in xlsx_qs if q["priority"] is not None}

    # Build PlaybookQuestions
    pb_questions: list[PlaybookQuestion] = []
    for i, rec in enumerate(docx_records, start=1):
        qid = rec["qid"]
        try:
            prio = int(qid[1:])
        except ValueError:
            prio = i
        xq = xlsx_by_prio.get(prio)

        pq = PlaybookQuestion(
            id=qid,
            priority=prio,
            condition_type=xq["condition_type"] if xq else "",
            question_text=rec["text"],
            extraction_hint=xq["extract"] if xq else None,
            output_type=_normalize_output(xq.get("output") if xq else None, "Text"),
            search_scope=_detect_search_scope(rec["text"]),
            keywords=[],
            yes_branch=_action_from_branch_text(rec.get("yes")),
            no_branch=_action_from_branch_text(rec.get("no")),
            red_flag=xq["red_flag"] if xq else None,
            notes="\n".join(rec.get("notes", [])) or None,
        )
        pb_questions.append(pq)

    # If docx had no parsable questions (e.g. Critical/Other clauses often have
    # just one yes/no), but xlsx has rows — synthesize one playbook question
    # per xlsx row.
    if not pb_questions and xlsx_qs:
        for xq in xlsx_qs:
            qid = f"Q{xq['priority'] or (len(pb_questions) + 1)}"
            pb_questions.append(PlaybookQuestion(
                id=qid,
                priority=xq["priority"] or (len(pb_questions) + 1),
                condition_type=xq["condition_type"],
                question_text=xq["question"],
                extraction_hint=xq["extract"],
                output_type=_normalize_output(xq["output"], "Text"),
                search_scope=SearchScope.ALL,
                yes_branch=PlaybookAction.extract_then(None),
                no_branch=PlaybookAction.record_none(),
                red_flag=xq["red_flag"],
            ))

    # Fallback: absolute minimum — a "does this clause exist?" question
    if not pb_questions:
        pb_questions.append(PlaybookQuestion(
            id="Q1",
            priority=1,
            condition_type="Definition Based",
            question_text=f"Does the lease contain any clause relating to {canonical_name}?",
            extraction_hint="Extract the entire clause as-is if present",
            output_type="Text",
            search_scope=SearchScope.ALL,
            yes_branch=PlaybookAction.extract_then(None),
            no_branch=PlaybookAction.record_none(),
        ))

    # Summary keywords — extract from docx overview where "Summary Section Keywords" line appears
    summary_keywords: list[str] = []

    pb = Playbook(
        field_id=field_id,
        field_name=canonical_name,
        category=category,
        output_type=qx["output_type_by_field"].get(canonical_name, "Text"),
        property_applicability=(
            _detect_property_restriction(docx_field_name)
            or qx["property_applicability"].get(canonical_name, {})
        ),
        source_docx=source_docx,
        overview=" ".join(docx_block.get("overview", [])).strip() or None,
        preliminary=preliminary,
        questions=pb_questions,
        keywords=qx["keywords_by_field"].get(canonical_name, []),
        summary_keywords=summary_keywords,
    )

    # Dependencies: Allowance depends on Lease Commencement Date (per docx)
    if canonical_name == "Allowance":
        pb.depends_on.append("original_lease_commencement_date")
    if canonical_name == "Lease Term (yrs.)":
        pb.depends_on.extend([
            "original_lease_commencement_date",
            "lease_expiration_date",
        ])

    return pb


# ---------------------------------------------------------------------------
# Main compile
# ---------------------------------------------------------------------------

def compile_all(source_dir: Path, out_dir: Path) -> dict:
    """Compile every playbook. Returns the index written to _index.json."""
    out_dir.mkdir(parents=True, exist_ok=True)

    qx = _load_questions_xlsx(source_dir / "Questions.xlsx")

    index: list[dict] = []

    for docx_name, category in DOCX_CATEGORY.items():
        docx_path = source_dir / docx_name
        if not docx_path.exists():
            print(f"  ! missing {docx_path}")
            continue
        blocks = _extract_field_blocks(docx_path)
        preliminary = blocks.pop("_preliminary", {}).get("lines", [])

        for docx_field_name, block in blocks.items():
            pb = _compile_one_playbook(
                docx_field_name, block, category, docx_name, qx, preliminary
            )
            if pb is None:
                continue
            out_path = out_dir / f"{pb.field_id}.json"
            with out_path.open("w") as fh:
                json.dump(_pb_to_dict(pb), fh, indent=2)
            index.append({
                "field_id": pb.field_id,
                "field_name": pb.field_name,
                "category": pb.category,
                "source_docx": pb.source_docx,
                "question_count": len(pb.questions),
                "keyword_count": len(pb.keywords),
                "output_type": pb.output_type,
                "file": f"{pb.field_id}.json",
            })

    # Fields that exist in Questions.xlsx but not in any docx — still compile
    # using just the xlsx metadata (Critical/Other often fall back here).
    seen_names = {e["field_name"] for e in index}
    for canonical_name, qs in qx["questions_by_field"].items():
        if canonical_name in seen_names:
            continue
        category = QXLSX_CATEGORY_MAP.get(qs[0]["main_field"], "Other Lease Clauses") if qs else "Other Lease Clauses"
        pb = _compile_one_playbook(
            canonical_name,  # reuse alias function below
            {"overview": [], "question_lines": [], "tables": []},
            category,
            "Questions.xlsx",
            qx,
            [],
        )
        # _compile_one_playbook expects a docx_field_name that maps to the alias.
        # Since Questions.xlsx already has canonical names, inject directly:
        if pb is None and canonical_name in qx["questions_by_field"]:
            field_id = _slug(canonical_name)
            qs_xl = qx["questions_by_field"][canonical_name]
            pb_questions = []
            for xq in qs_xl:
                qid = f"Q{xq['priority'] or (len(pb_questions) + 1)}"
                pb_questions.append(PlaybookQuestion(
                    id=qid,
                    priority=xq["priority"] or (len(pb_questions) + 1),
                    condition_type=xq["condition_type"],
                    question_text=xq["question"],
                    extraction_hint=xq["extract"],
                    output_type=_normalize_output(xq["output"], "Text"),
                    search_scope=SearchScope.ALL,
                    yes_branch=PlaybookAction.extract_then(None),
                    no_branch=PlaybookAction.record_none(),
                    red_flag=xq["red_flag"],
                ))
            pb = Playbook(
                field_id=field_id,
                field_name=canonical_name,
                category=category,
                output_type=qx["output_type_by_field"].get(canonical_name, "Text"),
                property_applicability=qx["property_applicability"].get(canonical_name, {}),
                source_docx="Questions.xlsx",
                questions=pb_questions,
                keywords=qx["keywords_by_field"].get(canonical_name, []),
            )
        if pb is None:
            continue
        out_path = out_dir / f"{pb.field_id}.json"
        with out_path.open("w") as fh:
            json.dump(_pb_to_dict(pb), fh, indent=2)
        index.append({
            "field_id": pb.field_id,
            "field_name": pb.field_name,
            "category": pb.category,
            "source_docx": pb.source_docx,
            "question_count": len(pb.questions),
            "keyword_count": len(pb.keywords),
            "output_type": pb.output_type,
            "file": f"{pb.field_id}.json",
        })

    # Write index
    index.sort(key=lambda e: (e["category"], e["field_name"]))
    with (out_dir / "_index.json").open("w") as fh:
        json.dump({"count": len(index), "playbooks": index}, fh, indent=2)

    return {"count": len(index), "playbooks": index}


def _pb_to_dict(pb: Playbook) -> dict:
    """Convert a Playbook to a JSON-serializable dict."""
    def _q_to_dict(q: PlaybookQuestion) -> dict:
        return {
            "id": q.id,
            "priority": q.priority,
            "condition_type": q.condition_type,
            "question_text": q.question_text,
            "extraction_hint": q.extraction_hint,
            "output_type": q.output_type,
            "search_scope": q.search_scope.value,
            "keywords": q.keywords,
            "yes_branch": _action_to_dict(q.yes_branch),
            "no_branch": _action_to_dict(q.no_branch),
            "red_flag": q.red_flag,
            "notes": q.notes,
        }

    def _action_to_dict(a: PlaybookAction | None) -> dict | None:
        if a is None:
            return None
        return {
            "type": a.type.value,
            "goto": a.goto,
            "literal": a.literal,
            "also_extract": a.also_extract,
        }

    return {
        "field_id": pb.field_id,
        "field_name": pb.field_name,
        "category": pb.category,
        "output_type": pb.output_type,
        "property_applicability": pb.property_applicability,
        "abstract_applicability": pb.abstract_applicability,
        "source_docx": pb.source_docx,
        "overview": pb.overview,
        "preliminary": pb.preliminary,
        "questions": [_q_to_dict(q) for q in pb.questions],
        "keywords": pb.keywords,
        "summary_keywords": pb.summary_keywords,
        "amendment_controls": pb.amendment_controls,
        "depends_on": pb.depends_on,
    }


if __name__ == "__main__":
    source = Path("data/playbooks_source")
    out = Path("data/playbooks_compiled")
    result = compile_all(source, out)
    print(f"Compiled {result['count']} playbooks into {out}")
    for e in result["playbooks"][:10]:
        print(f"  {e['category']:25s} {e['field_name']:40s} questions={e['question_count']:2d} output={e['output_type']}")
    if len(result["playbooks"]) > 10:
        print(f"  ... ({len(result['playbooks']) - 10} more)")
